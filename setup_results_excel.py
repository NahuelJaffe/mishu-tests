#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Configurar Excel Resultados para Testing Manual
Crear plantilla lista para documentar resultados de tests
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def setup_results_excel():
    # Crear nuevo workbook
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Crear hojas
    ws1 = wb.create_sheet("Resultados Tests")
    ws2 = wb.create_sheet("Instrucciones")
    ws3 = wb.create_sheet("Resumen KPIs")
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    title_font = Font(bold=True, color="2E86AB", size=14)
    subtitle_font = Font(bold=True, color="366092", size=11)
    normal_font = Font(size=10)
    pass_font = Font(bold=True, color="28A745")  # Verde
    fail_font = Font(bold=True, color="DC3545")  # Rojo
    pending_font = Font(bold=True, color="FFC107")  # Amarillo
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== HOJA 1: RESULTADOS TESTS ==========
    
    # Título principal
    ws1.merge_cells('A1:J1')
    ws1['A1'] = "RESULTADOS DE TESTING MANUAL - MISHU QA"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        "Test Scenario", "ID", "Test Case", "Fecha", "Tester", 
        "Estado", "Observaciones", "Evidencia", "Screenshot", "Notas Adicionales"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos de los Smoke Tests (5 tests)
    smoke_tests = [
        ["Smoke Test", "SMK-01", "Smoke – Login", "", "", "Pendiente", "", "", "", ""],
        ["Smoke Test", "SMK-02", "Smoke – Navegación principal", "", "", "Pendiente", "", "", "", ""],
        ["Smoke Test", "SMK-03", "Smoke – Connections list", "", "", "Pendiente", "", "", "", ""],
        ["Smoke Test", "SMK-04", "Smoke – Mensajes en conversación", "", "", "Pendiente", "", "", "", ""],
        ["Smoke Test", "SMK-05", "Smoke – Settings", "", "", "Pendiente", "", "", "", ""]
    ]
    
    # Escribir datos de Smoke Tests
    for row_num, test_case in enumerate(smoke_tests, 3):
        for col_num, value in enumerate(test_case, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Colorear según estado
            if col_num == 6:  # Columna Estado
                if value == "Pendiente":
                    cell.font = pending_font
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                elif value == "Pass":
                    cell.font = pass_font
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif value == "Fail":
                    cell.font = fail_font
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # Agregar validación de datos para Estado
    dv_state = DataValidation(type="list", formula1='"Pass,Fail,Skip,Blocked,Pendiente"')
    dv_state.add(f"F3:F{len(smoke_tests) + 2}")
    ws1.add_data_validation(dv_state)
    
    # Agregar validación de datos para Tester
    dv_tester = DataValidation(type="list", formula1='"Nahuel,QA Team,Automated"')
    dv_tester.add(f"E3:E{len(smoke_tests) + 2}")
    ws1.add_data_validation(dv_tester)
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 20,  # Test Scenario
        'B': 12,  # ID
        'C': 30,  # Test Case
        'D': 15,  # Fecha
        'E': 15,  # Tester
        'F': 12,  # Estado
        'G': 40,  # Observaciones
        'H': 20,  # Evidencia
        'I': 20,  # Screenshot
        'J': 30   # Notas Adicionales
    }
    
    for col_letter, width in column_widths.items():
        ws1.column_dimensions[col_letter].width = width
    
    # Agregar fórmulas para fecha automática
    for row in range(3, len(smoke_tests) + 3):
        ws1.cell(row=row, column=4, value="=TODAY()")  # Fecha automática
    
    # ========== HOJA 2: INSTRUCCIONES ==========
    
    # Título principal
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "INSTRUCCIONES PARA USAR ESTE EXCEL"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    # Contenido de instrucciones
    instructions = [
        ["CÓMO USAR ESTE EXCEL", "", "", ""],
        ["1. Ejecutar Test Manual", "", "", ""],
        ["   - Abrir la aplicación en el navegador", "https://mishu-web--pr69-performance-and-prof-8fsc02so.web.app/", "", ""],
        ["   - Seguir los pasos del test case", "Ver Excel v3 - All Test Cases", "", ""],
        ["   - Documentar resultado en esta hoja", "Columna 'Estado'", "", ""],
        ["", "", "", ""],
        ["2. Llenar Información", "", "", ""],
        ["   - Tester: Tu nombre", "Nahuel", "", ""],
        ["   - Estado: Pass/Fail/Skip/Blocked", "Usar lista desplegable", "", ""],
        ["   - Observaciones: Qué pasó", "Detalles del resultado", "", ""],
        ["   - Evidencia: Link o referencia", "URL, ticket, etc.", "", ""],
        ["   - Screenshot: Nombre del archivo", "screenshot_smk01_pass.png", "", ""],
        ["", "", "", ""],
        ["3. Estados Posibles", "", "", ""],
        ["   Pass ✅", "Test funcionó correctamente", "Verde", "D4EDDA"],
        ["   Fail ❌", "Test falló o encontró bug", "Rojo", "F8D7DA"],
        ["   Skip ⏭️", "Test saltado (no aplica)", "Amarillo", "FFF3CD"],
        ["   Blocked 🚫", "Test bloqueado por otro bug", "Gris", "E2E3E5"],
        ["   Pendiente ⏳", "Test no ejecutado aún", "Amarillo", "FFF3CD"],
        ["", "", "", ""],
        ["4. Ejemplos de Observaciones", "", "", ""],
        ["   Pass: 'Login exitoso, redirige a /connections'", "", "", ""],
        ["   Fail: 'Error 500 al hacer login, mensaje: Internal Server Error'", "", "", ""],
        ["   Skip: 'Test no aplica en esta versión'", "", "", ""],
        ["   Blocked: 'No se puede probar por bug en SMK-01'", "", "", ""],
        ["", "", "", ""],
        ["5. Screenshots", "", "", ""],
        ["   - Tomar screenshot cuando encuentres bugs", "Cmd+Shift+4 (Mac) o Snipping Tool (Windows)", "", ""],
        ["   - Guardar con nombre descriptivo", "screenshot_smk01_fail_login_error.png", "", ""],
        ["   - Referenciar en columna 'Screenshot'", "Nombre del archivo", "", ""],
        ["", "", "", ""],
        ["6. Buenas Prácticas", "", "", ""],
        ["   - Ser específico en observaciones", "No solo 'no funciona'", "", ""],
        ["   - Incluir pasos para reproducir bugs", "1. Hacer login 2. Ver error", "", ""],
        ["   - Documentar fecha y hora", "Automático con fórmula", "", ""],
        ["   - Revisar antes de marcar como Pass", "Verificar que realmente funciona", "", ""]
    ]
    
    for row_num, instruction in enumerate(instructions, 2):
        for col_num, value in enumerate(instruction, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            if row_num == 2 or row_num == 7 or row_num == 14 or row_num == 21 or row_num == 26 or row_num == 31:
                cell.font = subtitle_font
                cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            else:
                cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajustar ancho de columnas
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 15
    
    # ========== HOJA 3: RESUMEN KPIS ==========
    
    # Título principal
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "RESUMEN Y KPIS DE TESTING"
    ws3['A1'].font = title_font
    ws3['A1'].alignment = Alignment(horizontal='center')
    
    # KPIs y métricas
    kpis = [
        ["MÉTRICAS GENERALES", "", "", ""],
        ["Total Tests Ejecutados", "=COUNTA(Resultados!F3:F7)", "Fórmula automática", ""],
        ["Tests Pass", "=COUNTIF(Resultados!F3:F7,\"Pass\")", "Fórmula automática", ""],
        ["Tests Fail", "=COUNTIF(Resultados!F3:F7,\"Fail\")", "Fórmula automática", ""],
        ["Tests Skip", "=COUNTIF(Resultados!F3:F7,\"Skip\")", "Fórmula automática", ""],
        ["Tests Blocked", "=COUNTIF(Resultados!F3:F7,\"Blocked\")", "Fórmula automática", ""],
        ["Tests Pendientes", "=COUNTIF(Resultados!F3:F7,\"Pendiente\")", "Fórmula automática", ""],
        ["", "", "", ""],
        ["PORCENTAJES", "", "", ""],
        ["% Pass Rate", "=B4/B3*100", "Fórmula automática", "%"],
        ["% Fail Rate", "=B5/B3*100", "Fórmula automática", "%"],
        ["% Completion", "=(B3-B8)/B3*100", "Fórmula automática", "%"],
        ["", "", "", ""],
        ["ESTADO POR TEST", "", "", ""],
        ["SMK-01", "=Resultados!F3", "Referencia automática", ""],
        ["SMK-02", "=Resultados!F4", "Referencia automática", ""],
        ["SMK-03", "=Resultados!F5", "Referencia automática", ""],
        ["SMK-04", "=Resultados!F6", "Referencia automática", ""],
        ["SMK-05", "=Resultados!F7", "Referencia automática", ""],
        ["", "", "", ""],
        ["FECHAS", "", "", ""],
        ["Fecha de Testing", "=TODAY()", "Fórmula automática", ""],
        ["Última Actualización", "=NOW()", "Fórmula automática", ""],
        ["", "", "", ""],
        ["NOTAS", "", "", ""],
        ["Bugs Encontrados", "Contar manualmente", "Revisar columna Observaciones", ""],
        ["Tests Críticos", "SMK-01, SMK-02", "Login y navegación", ""],
        ["Próximos Pasos", "Automatizar tests que pasan", "Después de testing manual", ""]
    ]
    
    for row_num, kpi in enumerate(kpis, 2):
        for col_num, value in enumerate(kpi, 1):
            cell = ws3.cell(row=row_num, column=col_num, value=value)
            if row_num == 2 or row_num == 9 or row_num == 14 or row_num == 21 or row_num == 25:
                cell.font = subtitle_font
                cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            else:
                cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajustar ancho de columnas
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['D'].width = 15
    
    # Guardar archivo
    excel_file = '/Users/nahueljaffe/Desktop/Resultados_Testing_Manual_Mishu.xlsx'
    wb.save(excel_file)
    
    print(f"Excel Resultados configurado exitosamente: {excel_file}")
    print(f"Total de hojas: 3")
    print(f"Smoke Tests configurados: 5")
    print(f"Fórmulas automáticas: KPIs y fechas")
    
    return excel_file

if __name__ == "__main__":
    try:
        excel_file = setup_results_excel()
        print(f"\nArchivo listo para usar: {excel_file}")
        print(f"\nContenido del Excel:")
        print(f"  - Hoja 1: Resultados Tests (5 Smoke Tests)")
        print(f"  - Hoja 2: Instrucciones detalladas")
        print(f"  - Hoja 3: Resumen KPIs automático")
        print(f"\nCaracterísticas:")
        print(f"  - Listas desplegables para Estado y Tester")
        print(f"  - Fórmulas automáticas para fechas y KPIs")
        print(f"  - Colores automáticos según estado")
        print(f"  - Instrucciones paso a paso")
        print(f"\n¡Perfecto para documentar testing manual!")
    except Exception as e:
        print(f"Error: {e}")
